import tkinter as tk
from tkinter import filedialog,ttk,messagebox
import pandas as pd
import calendar
from datetime import datetime, timedelta
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Alignment,PatternFill,Border, Side,Font
from openpyxl.worksheet.dimensions import ColumnDimension,DimensionHolder
from openpyxl.utils import get_column_letter
# from win32com.client import Dispatch
import os

root = tk.Tk()
root.geometry("600x300+50+50")
root.title("工程/客服/保安考勤记录生成器")

select_path = tk.StringVar()
select_path_lastmonth = tk.StringVar()

def select_file():
    selected_file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")])
    select_path.set(selected_file_path)

def select_file_lastmonth():
    selected_file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")])
    select_path_lastmonth.set(selected_file_path)
    # print(get_remaining_hours("李琦琛"))
   
def get_remaining_hours(name) -> list:
    data_frame = pd.read_excel(select_path_lastmonth.get())
     
    # 初始化行列位置
    row = 1
    column = None
    result_list = ["","","",""]

    # 遍历DataFrame寻找内容
    for j in range(len(data_frame.columns)):
        if data_frame.iloc[row, j] == name:
            column = j
            # print(f'内容找到在行：{row+1}，列：{j+1}')
            break

    # 如果需要获取具体的单元格数据
    if column is not None:
        row_idx = data_frame.shape[0] - 4
   
        for list_index in range(0,4):
            if not pd.isna(data_frame.iloc[row_idx, column + list_index]):
                last_cell_data = float(data_frame.iloc[row_idx, column + list_index])
                if not pd.isna(data_frame.iloc[row_idx + 1, column]):
                    cur_cell_data = float(data_frame.iloc[row_idx + 1, column + list_index])
                    result_list[list_index] = str(cur_cell_data + last_cell_data)
                else:
                    result_list[list_index] = str(last_cell_data)
            else:
                if not pd.isna(data_frame.iloc[row_idx + 1, column + list_index]):
                    cur_cell_data = float(data_frame.iloc[row_idx + 1, column + list_index])
                    result_list[list_index] = str(cur_cell_data)
                else:
                    result_list[list_index] = ''
    # print(result_list)
    return result_list

def generate_excel():
    try:
        selected_month = month_combo.get().replace('月','')
        selected_year = year_combo.get()
        # print(get_days_of_month(int(selected_year), int(selected_month)))
        workbook = Workbook()
        sheet = workbook.active

        # 创建实线边框的边界定义
        border = Border(left=Side(border_style='thin', color='000000'),  # 左边界，实线，颜色为黑色
                    right=Side(border_style='thin', color='000000'),  # 右边界
                    top=Side(border_style='thin', color='000000'),  # 上边界
                    bottom=Side(border_style='thin', color='000000'))  # 下边界

        
        sheet.merge_cells('A3:A4')
        sheet.cell(row=3, column=1).value = '星期'
        sheet.cell(row=3, column=1).alignment = Alignment(horizontal="center",vertical="center")
        sheet.cell(row=3, column=1).border = border

        sheet.cell(row=3, column=2).value = '姓名'
        sheet.cell(row=3, column=2).alignment = Alignment(horizontal="center",vertical="center")
        sheet.cell(row=3, column=2).border = border

        sheet.cell(row=4, column=2).value = '日期'
        sheet.cell(row=4, column=2).alignment = Alignment(horizontal="center",vertical="center")
        sheet.cell(row=4, column=2).border = border

        df = pd.read_excel(select_path.get())
        print(df)
        
        # 生成日期数据
        for column_index in range(1,32):
            date_iloc = df.iloc[1, column_index]  #取日期数据
            if type(date_iloc).__name__ == "int":
                row_index = column_index + 4
                
                # print(date_iloc)
                sheet.cell(row=row_index, column=2).value = f'{selected_month}月{date_iloc}日'
                sheet.cell(row=row_index, column=2).alignment = Alignment(horizontal="center",vertical="center")
                sheet.cell(row=row_index, column=2).border = border

                day_iloc = df.iloc[2, column_index]  #取星期几数据
                sheet.cell(row=row_index, column=1).value = day_iloc
                sheet.cell(row=row_index, column=1).alignment = Alignment(horizontal="center",vertical="center")
                sheet.cell(row=row_index, column=1).border = border

                if day_iloc == '六' or day_iloc == '日':
                    for col in range(1,sheet.max_column + 1):
                        sheet.cell(row=row_index, column=col).fill = PatternFill(start_color='FFFF00', fill_type='solid')
        
        #生成最后几行汇总数据的第一列内容
        row_index = row_index + 1
        sheet.cell(row=row_index, column=1).value = f'{selected_month}月合计'
        sheet.cell(row=row_index, column=1).alignment = Alignment(horizontal="center",vertical="center")
        sheet.cell(row=row_index, column=1).border = border
        sheet.cell(row=row_index, column=1).font = Font(size=10, bold=True)
        sheet.merge_cells(f'A{row_index}:B{row_index}')

        row_index = row_index + 1
        sheet.cell(row=row_index, column=1).value = f'{int(selected_month) - 1}月剩余加班小时'
        sheet.cell(row=row_index, column=1).alignment = Alignment(horizontal="center",vertical="center")
        sheet.cell(row=row_index, column=1).border = border
        sheet.cell(row=row_index, column=1).font = Font(size=8, bold=True)
        sheet.merge_cells(f'A{row_index}:B{row_index}')

        row_index = row_index + 1
        sheet.cell(row=row_index, column=1).value = f'{selected_month}月剩余加班小时数'
        sheet.cell(row=row_index, column=1).alignment = Alignment(horizontal="center",vertical="center")
        sheet.cell(row=row_index, column=1).border = border
        sheet.cell(row=row_index, column=1).font = Font(size=8, bold=True)
        sheet.merge_cells(f'A{row_index}:B{row_index}')

        row_index = row_index + 1
        sheet.cell(row=row_index, column=1).value = "签名"
        sheet.cell(row=row_index, column=1).alignment = Alignment(horizontal="center",vertical="center")
        sheet.cell(row=row_index, column=1).border = border
        sheet.cell(row=row_index, column=1).font = Font(size=11, bold=True)
        sheet.merge_cells(f'A{row_index}:B{row_index}')
        sheet.row_dimensions[row_index].height = 35

        row_index = row_index + 1
        sheet.cell(row=row_index, column=1).value = "制表人："
        sheet.cell(row=row_index, column=1).alignment = Alignment(horizontal="center",vertical="bottom")
        sheet.merge_cells(f'A{row_index}:B{row_index}')

        sheet.cell(row=row_index, column=8).value = "项目负责人："
        sheet.cell(row=row_index, column=8).alignment = Alignment(horizontal="left",vertical="bottom")
        sheet.row_dimensions[row_index].height = 31.1

        #生成员工数据
        name_row_index = 3
        name_column_index = 2
    
        while True:
            name_iloc = df.iloc[name_row_index, 0]
            if name_iloc == "排班说明":
                break
            else: # 一个员工的数据
                sheet.cell(row=3, column=name_column_index + 1).value = name_iloc
                sheet.cell(row=3, column=name_column_index + 1).alignment = Alignment(horizontal="center",vertical="center")
                sheet.cell(row=3, column=name_column_index + 1).border = border
                sheet.cell(row=3, column=name_column_index + 1).font = Font(size=11)
                # 合并“姓名”的单元格
                sheet.merge_cells(start_row=3, start_column=name_column_index + 1, end_row=3, end_column=name_column_index + 4)

                last_month_hours = get_remaining_hours(name_iloc) # 获取上个月的剩余加班小时数

                cell1 = sheet.cell(row=4, column=name_column_index + 1)
                cell1.value = '平时加班'
                cell1.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                cell1.border = border
                cell1.font = Font(size=8)
                sheet.column_dimensions[get_column_letter(cell1.column)].width = 4.2
                set_cell_border(row_index,sheet,name_column_index + 1,border)
                cell_total_1 = sheet.cell(row=row_index - 4, column=name_column_index + 1) 
                cell_total_1.value = f'=SUM({cell1.column_letter}5:{cell1.column_letter}{row_index - 5})'
                # 写入上个月的平时加班小时数
                ordinary_ot_cell = sheet.cell(row=row_index - 3, column=name_column_index + 1)
                ordinary_ot_cell.value = last_month_hours[0]
    

                cell2 = sheet.cell(row=4, column=name_column_index + 2)
                cell2.value = '法定加班'
                cell2.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                cell2.border = border
                cell2.font = Font(size=8)
                sheet.column_dimensions[get_column_letter(cell2.column)].width = 4.2
                set_cell_border(row_index,sheet,name_column_index + 2,border)
                cell_total_2 = sheet.cell(row=row_index - 4, column=name_column_index + 2) 
                cell_total_2.value = f'=SUM({cell2.column_letter}5:{cell2.column_letter}{row_index - 5})'
                # 写入上个月的法定加班小时数
                statutory_ot_cell = sheet.cell(row=row_index - 3, column=name_column_index + 2)
                statutory_ot_cell.value = last_month_hours[1]
    

                cell3 = sheet.cell(row=4, column=name_column_index + 3)
                cell3.value = '周末加班'
                cell3.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                cell3.border = border
                cell3.font = Font(size=8)
                sheet.column_dimensions[get_column_letter(cell3.column)].width = 4.2
                set_cell_border(row_index,sheet,name_column_index + 3,border)
                cell_total_3 = sheet.cell(row=row_index - 4, column=name_column_index + 3) 
                cell_total_3.value = f'=SUM({cell3.column_letter}5:{cell3.column_letter}{row_index - 5})'
                # 写入上个月的周末加班小时数
                weekend_ot_cell = sheet.cell(row=row_index - 3, column=name_column_index + 3)
                weekend_ot_cell.value = last_month_hours[2]


                cell4 = sheet.cell(row=4, column=name_column_index + 4)
                cell4.value = '调休'
                cell4.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                cell4.border = border
                cell4.font = Font(size=8)
                sheet.column_dimensions[get_column_letter(cell4.column)].width = 4.2
                set_cell_border(row_index,sheet,name_column_index + 4,border)
                cell_total_4 = sheet.cell(row=row_index - 4, column=name_column_index + 4) 
                cell_total_4.value = f'=SUM({cell4.column_letter}5:{cell4.column_letter}{row_index - 5})'
            
                # 合并“签名”的单元格
                sheet.cell(row=row_index - 1, column=name_column_index + 1).border = border
                sheet.merge_cells(start_row=row_index - 1, start_column=name_column_index + 1, end_row=row_index - 1, end_column=name_column_index + 4)

                rest_hours = 0 #一个员工当月的调休小时数
                for date_col_index in range(1,32):
                    result_iloc = df.iloc[name_row_index, date_col_index]
                    if result_iloc == "调休":
                        name_value = df.iloc[name_row_index, 0]
                        date_value = df.iloc[1, date_col_index]
                        # day_value = df.iloc[2, date_col_index]
                        for cell in sheet[3]:
                            if cell.value == name_value:
                                sheet.cell(row=4 + date_value, column=cell.column + 3).value = 8
                                rest_hours = rest_hours + 8
                                    
                    if result_iloc == "加班":
                        name_value = df.iloc[name_row_index, 0]
                        date_value = df.iloc[1, date_col_index]
                        day_value = df.iloc[2, date_col_index] 
                
                        for cell in sheet[3]:
                            if cell.value == name_value:
                                if day_value == "六" or day_value == "日":
                                    sheet.cell(row=4 + date_value, column=cell.column + 2).value = 8
                                else:
                                    sheet.cell(row=4 + date_value, column=cell.column).value = 8

    
                name_column_index = name_column_index + 4
                name_row_index = name_row_index + 1

        selected_type = type_combo.get()
        content = f"{selected_type}{selected_year}年{selected_month}月加班调休明细表"
        sheet.cell(row=1, column=1).value = content
        sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center",vertical="center")
        sheet.cell(row=1, column=1).font = Font(size=18)
        sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=sheet.max_column)
        global file_name
        file_name = f"{content}.xlsx"
        workbook.save(file_name)

        messagebox.showinfo("提示", "生成明细表成功")
        button2.config(state=tk.ACTIVE)
    except Exception as e:
        messagebox.showerror("错误", "生成明细表失败，请检查选择的文件内容是否正确!原因：" + repr(e))

    finally:
        workbook.close()

def recalculate_left_hours():
    try:
        '''
                root = os.getcwd()   
        # 需要先打开保存一遍Excel文件才能不到有公式的单元格
        xlApp = Dispatch("Excel.Application")
        xlApp.Visible = False
        xlBook = xlApp.Workbooks.Open(os.path.join(root, file_name))
        xlBook.Save()
        xlBook.Close()
        '''


        # 读Excel文件用来取数据
        workbook = load_workbook(file_name,read_only=True,data_only=True)
        sheet = workbook.active
        # workbook.close()

        # 读Excel文件用来写数据
        write_workbook = load_workbook(file_name)
        write_sheet = write_workbook.active

        # print(sheet.max_row)
        # print(sheet.max_column)
        # 把本月剩余的加班小时数抄过来
        continue_count = 0
        for col_iter_index in range(3,sheet.max_column + 1):
            continue_count = continue_count + 1
            if continue_count % 4 == 0:
                continue
            current_total_hours = sheet.cell(row = sheet.max_row - 4,column = col_iter_index).value
            # print(f'********{current_total_hours}******')
            if current_total_hours is not None and float(current_total_hours) > 0:
                write_sheet.cell(row = write_sheet.max_row - 2,column = col_iter_index).value = current_total_hours
        
        # 计算每个员工的剩余加班小时数
        employees_count = (sheet.max_column - 3) / 4 # 员工数量
        start_col_index = 3 # 从第三列开始
        for _ in range(0,int(employees_count)):
            hours_data = list() # 员工的小时数数据
            for j in range(0,4):
                earch_hour = sheet.cell(row = sheet.max_row - 3,column = start_col_index + j).value
                hours_data.append(earch_hour if earch_hour is not None else "")
                if j == 3:
                    rest_hour = sheet.cell(row = sheet.max_row - 4,column = start_col_index + j).value
                    hours_data.append(rest_hour if rest_hour is not None else "")
            
            # 根据调休时间计算一个员工的剩余加班时间
            hours_data = cal_remaining_hours(0,hours_data)
            print(hours_data)
   
            if float(hours_data[4]) > 0: # 上个月剩余的加班小时数不够扣调休小时数
                for l in range(0,4):
                    earch_hour = sheet.cell(row = sheet.max_row - 4,column = start_col_index + l).value
                    hours_data[l] = earch_hour if earch_hour is not None else ""
                hours_data = cal_remaining_hours(0,hours_data) # 用本月的加班小时数扣调休小时数
                for n in range(0,3):
                    if len(hours_data[n]) > 0 and float(hours_data[n]) > 0:
                        write_sheet.cell(row = write_sheet.max_row - 2,column = start_col_index + n).value = str(hours_data[n])
            else: # 上个月剩余的加班小时数够扣调休小时数,直接更新本月剩余加班小时数
                for m in range(0,3):
                    write_sheet.cell(row = write_sheet.max_row - 3,column = start_col_index + m).value = str(hours_data[m])
            

            start_col_index = start_col_index + 4

        write_workbook.save(file_name)
        messagebox.showinfo("提示", "计算剩余小时数成功")

    except Exception as e:
        messagebox.showerror("错误", "计算剩余小时数失败，原因：" + repr(e))
        # raise e
    finally:
        write_workbook.close()

'''
 递归用剩余加班小时数减去调休小时数

last_remaining_hours的前4个元素是上个月或者本月的加班小时数，第5个元素是调休小时数
'''
def cal_remaining_hours(index,last_remaining_hours) -> list:
    if index <= 3:
        if last_remaining_hours[index] != "":
            if float(last_remaining_hours[index]) >= float(last_remaining_hours[4]):
                last_remaining_hours[index] = str(float(last_remaining_hours[index]) - float(last_remaining_hours[4]))
                last_remaining_hours[4] = '0'
            else:
                last_remaining_hours[4] = str(float(last_remaining_hours[4]) - float(last_remaining_hours[index]))
                last_remaining_hours[index] = '0'
                cal_remaining_hours(index + 1,last_remaining_hours)
             
        else:
            cal_remaining_hours(index + 1,last_remaining_hours)

    return last_remaining_hours

# 每个员工具体每一天的单元格设置边框
def set_cell_border(row_index,sheet,column_index,border):
    for border_row in range(5,row_index - 1):
        cell_border = sheet.cell(row=border_row, column=column_index)
        cell_border.alignment = Alignment(horizontal="center",vertical="center")
        cell_border.border = border
        cell_border.font = Font(size=9, bold=True)
        cell0 = sheet.cell(row=border_row, column=1)
        if cell0.value == "六" or cell0.value == "日":
            cell_border.fill = PatternFill(start_color='FFFF00', fill_type='solid')
  

def select_date_range():
    print(f"Selected month: {month_combo.get()}")

def get_days_of_month(year, month):
    # 获取指定年月的日历
    cal = calendar.monthcalendar(year, month)
    days_of_month = []
    for week in cal:
        for day in week:
            if day > 0:
                # 将日期转换为字符串格式
                day_str = str(day)
                # 获取星期几
                weekday = datetime(year, month, day).weekday()
                # 将星期几转换为星期几的名称
                weekday_name = calendar.day_name[weekday]
                # 打包成字典
                day_info = {
                    'day': day_str,
                    'weekday': weekday_name
                }
                days_of_month.append(day_info)
    return days_of_month

if __name__ == '__main__':
   
    years = list(range(2024,2055))
    year_combo = ttk.Combobox(root, values=years,width=5)
    year_combo.current(0)  # 设置默认选择为"一月"
    year_combo.configure(state="readonly")
    year_combo.grid(column=0, row=0)

    months = ["1月", "2月", "3月", "4月", "5月", "6月",
          "7月", "8月", "9月", "10月", "11月", "12月"]
    month_combo = ttk.Combobox(root, values=months,width=5)
    month_combo.current(0)  # 设置默认选择为"一月"
    month_combo.configure(state="readonly")
    month_combo.grid(column=1, row=0)
    
    types = ["工程", "客服", "保安"]
    type_combo = ttk.Combobox(root, values=types,width=5)
    type_combo.current(0)  
    type_combo.configure(state="readonly")
    type_combo.grid(column=2, row=0)

    entry = tk.Entry(root, textvariable=select_path,width=45)
    entry.grid(column=0, row=1)
    entry.configure(state="readonly")
    tk.Button(root, text="选择本月排班表", command=select_file).grid(row=1, column=1)

    entry = tk.Entry(root, textvariable=select_path_lastmonth,width=45)
    entry.grid(column=0, row=2)
    entry.configure(state="readonly")
    tk.Button(root, text="选择上个月加班调休明细表", command=select_file_lastmonth).grid(row=2, column=1)

    button1 = tk.Button(root, text="1、生成加班调休明细表",command=generate_excel)
    button1.grid(row=5, column=1, sticky="EWNS",pady=20)  # 使Button在row=1, column=1的位置，sticky选项使其在水平和垂直方向上扩展
    
    button2 = tk.Button(root, text="2、计算剩余加班小时数",command=recalculate_left_hours)
    button2.grid(row=6, column=1, sticky="EWNS",pady=10)
    button2.config(state=tk.DISABLED)


    root.mainloop()
 

